#-*- coding: utf-8 -*-
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re, os
import multiprocessing

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
path = BASE_DIR + "/down/chromedriver.exe"

blankRemovePattern = re.compile(r'\s+')
mapExtractPattern = re.compile(r'\((.*?)\)')
siteExtractPattern = re.compile(r'\'(.*?)\'')

ogNameDic = {}

def getList(indexId):
    urlPath = "http://e-childschoolinfo.moe.go.kr/kinderMt/combineFind.do?pageIndex=%d" % indexId
    driver.get(urlPath)
    driver.implicitly_wait(10)

    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    resultCount = soup.select_one('#noticeSearch > div.content > p.tblResult > span')
    print(resultCount.text.strip())

    resultTables = soup.select('#noticeSearch > div > table > tbody > tr')

    lineString = '================================================================='

    for idx, tr in enumerate(resultTables):
        # 엑셀 행값
        rowIndex = (indexId - 1) * 10 + idx

        print(lineString)

        tdObjects = tr.select('td');
        for idx, td in enumerate(tdObjects):
            if idx < 1:
                continue
            elif idx == 1:
                print('기관 유형 : %s' % td.text.strip())
                sheet1.cell(row=rowIndex, column=1).value = td.text.strip()
            elif idx == 2:
                value = td.select_one('a')
                matchString = mapExtractPattern.search(value['href']).group();
                matchString = matchString[1: len(matchString) - 1]
                splitMatchString = matchString.split(',');

                # 기관 아이디
                ogId = splitMatchString[1].strip()[1: len(splitMatchString[1].strip()) - 1]
                # 기관 유형 코드 : 01 - 유치원, 02 - 어린이집
                ogType =  splitMatchString[2].strip()[1: len(splitMatchString[2].strip()) - 1]
                # 기관명
                ogName = re.sub(blankRemovePattern, " ", value.text.strip())

                print('기관명 : %s' % re.sub(blankRemovePattern, " ", value.text.strip()))
                print('기관코드 : %s' % splitMatchString[1].strip()[1: len(splitMatchString[1].strip()) - 1])
                print('기관유형 : %s' % splitMatchString[2].strip()[1: len(splitMatchString[2].strip()) - 1])
                sheet1.cell(row=rowIndex, column=2).value = ogId
                sheet1.cell(row=rowIndex, column=3).value = ogName
                sheet1.cell(row=rowIndex, column=4).value = ogType

                # 유치원인 경우에 기관 정보를 입력
                if ogType == '01':
                    ogNameDic[ogId] = urlPath

            elif idx == 3:
                print('설립 유형 : %s' % td.text.strip())
                sheet1.cell(row=rowIndex, column=5).value = td.text.strip()
            elif idx == 4:
                print('기관 주소 : %s' % td.text.strip())
                sheet1.cell(row=rowIndex, column=6).value = td.text.strip()
            elif idx == 5:
                # 위도 : Latitude, 경도 : Longitude
                value = td.select_one('a')
                matchString = mapExtractPattern.search(value['href']).group();
                matchString = matchString[1: len(matchString) - 1]
                splitMatchString = matchString.split(',');
                print('Latitude : %s, Longitude : %s' % (splitMatchString[1].strip(), splitMatchString[2].strip()))
                sheet1.cell(row=rowIndex, column=7).value = splitMatchString[1].strip()
                sheet1.cell(row=rowIndex, column=8).value = splitMatchString[2].strip()
            elif idx == 6:
                value = td.select_one('a')
                if value is None:
                    continue
                else:
                    matchString = siteExtractPattern.search(value['href']).group();
                    matchString = matchString[1: len(matchString) - 1]
                    print('사이트 주소 : %s' % matchString)
                    sheet1.cell(row=rowIndex, column=9).value = matchString

        print(lineString)

def getDetailInfo(key, urlPath):
    driver.get(urlPath)
    driver.implicitly_wait(10)

    elem = driver.find_element_by_css_selector("[href*='%s']" % key)
    elem.click()

if __name__ == "__main__":
    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument("disable-gpu")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.80 Safari/537.36")

    # 혹은 options.add_argument("--disable-gpu")
    driver = webdriver.Chrome(path, chrome_options=options)

    wb = Workbook()
    sheet1 = wb.active

    fileName = '20181211_APP_BACK_DATA.xlsx'
    sheet1.title = '유치원어린이집목록'

    for idx in range(1, 2):
        getList(idx)

    wb.save(filename=fileName)

    for key, urlPath in ogNameDic.items():
        getDetailInfo(key, urlPath)

    driver.close()

    print(ogNameDic)
    print("MultiProcessing CPU Count : %d" % multiprocessing.cpu_count())
